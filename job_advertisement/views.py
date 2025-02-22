from django.shortcuts import render, redirect
from django.views import View
from job_advertisement import models
from datetime import datetime, timedelta
from django.contrib import messages
from client_management.models import AddCompanyModel
from django.db.models import Q
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.http import JsonResponse
from django.contrib.auth.models import User
import openpyxl


@method_decorator(login_required(login_url="/"), name='dispatch')
class CreateJobView(View):
    template_name = 'create-job.html'
    def get(self, request, id=None):
        datas = {'companies': AddCompanyModel.objects.all()}
        if id is not None:
            job = models.JobAdvertisementModel.objects.get(id=id)
            datas['job'] = job
            try:
                datas['date_of_app'] = job.date_of_app.strftime('%Y-%m-%d')
            except: pass
            try:
                datas['date_of_review'] = job.date_of_review.strftime('%Y-%m-%d')
            except: pass
        return render(request, self.template_name, context=datas)

    def post(self, request, id=None):
        do_app = request.POST.get('do-app')
        job_code = request.POST.get('job-code')
        no_of_vacancy = request.POST.get('no-vacany')
        range_field = request.POST.get('range')
        designation = request.POST.get('designation')
        comp_name = request.POST.get('comp-name')
        dor = request.POST.get('dor')
        no_app = request.POST.get('no_app')

        try:
            no_app = int(no_app)
        except:
            no_app = 0

        try:
            dor = datetime.strptime(dor, '%Y-%m-%d')
        except:
            dor = None

        try:
            do_app = datetime.strptime(do_app, '%Y-%m-%d')
            maturity_date = do_app + timedelta(days=15)
            expiry_date = maturity_date + timedelta(days=90)
        except Exception as e:
            messages.error(request, f'Error: {e}')

        try:
            comp_name = models.AddCompanyModel.objects.get(id=comp_name)
        except Exception:
            comp_name = None

        if models.JobAdvertisementModel.objects.filter(job_code=job_code) and id is None:
            messages.error(request, f'{job_code} Already Exists!!')
            return redirect('/job-advertisement/create-job')

        if id is not None:
            try:
                job = models.JobAdvertisementModel.objects.get(id=id)
                job.date_of_app = do_app
                job.no_of_vaccancies = no_of_vacancy
                job.maturity_date = maturity_date
                job.expiry_date = expiry_date
                job.job_code = job_code
                job.range_field = range_field
                job.designation = designation
                job.comp_name = comp_name
                job.date_of_review = dor
                job.no_of_app = no_app
                job.save()
                messages.success(request, f'{job_code} Updated..')
            except Exception as e:
                messages.error(request, f'Error: {e}')
        else:
            try:
                job = models.JobAdvertisementModel(date_of_app=do_app, maturity_date=maturity_date, expiry_date=expiry_date,
                        job_code=job_code, range_field=range_field, designation=designation, comp_name=comp_name,
                        no_of_vaccancies=no_of_vacancy, date_of_review=dor, no_of_app=no_app)
                job.save()
                messages.success(request, f'{job_code} New Job Advertisement Created..')
            except Exception as e:
                messages.error(request, f'Error: {e}')
        return redirect('/job-advertisement/create-job')

@method_decorator(login_required(login_url="/"), name='dispatch')
class JobListView(View):
    template_name = 'job-list.html'
    def get(self, request, other_url=None):
        keyword = request.GET.get('keyword')
        get_url = request.get_full_path()

        if '?keyword' in get_url:
            get_url = get_url.split('&page=')[0]
            current_url = f"{get_url}&"
        else:
            get_url = get_url.split('?')[0]
            current_url = f"{get_url}?"

        today = datetime.today().date()

        if keyword:
            try:
                keyword = datetime.strptime(keyword, '%d/%m/%Y').date()
            except: pass
            if other_url is None:
                jobs = models.JobAdvertisementModel.objects.filter(
                        Q(date_of_app__icontains=keyword)|Q(job_code__icontains=keyword)|
                        Q(maturity_date__icontains=keyword)|Q(expiry_date__icontains=keyword)|
                        Q(designation__icontains=keyword)|Q(no_of_vaccancies__icontains=keyword)|
                        Q(no_of_app__icontains=keyword)|Q(comp_name__company_name__icontains=keyword)
                        ).filter(expiry_date__gte=today).order_by('-date_of_app')
            else:
                jobs = models.JobAdvertisementModel.objects.filter(
                        Q(date_of_app__icontains=keyword)|Q(job_code__icontains=keyword)|
                        Q(maturity_date__icontains=keyword)|Q(expiry_date__icontains=keyword)|
                        Q(designation__icontains=keyword)|Q(no_of_vaccancies__icontains=keyword)|
                        Q(no_of_app__icontains=keyword)|Q(comp_name__company_name__icontains=keyword)
                        ).order_by('-date_of_app')
            result_cnt = jobs.count()

        if other_url == 'matured-jobs':
            if keyword:
                jobs = jobs.filter(maturity_date__lte=today, expiry_date__gte=today,
                                    date_of_review__isnull=True).order_by('-maturity_date')
            else:
                jobs = models.JobAdvertisementModel.objects.filter(maturity_date__lte=today, expiry_date__gte=today,
                                    date_of_review__isnull=True).order_by('-maturity_date')
            result_cnt = jobs.count()
        elif other_url == 'remainder-jobs':
            if keyword:
                matured_jobs = jobs.filter(expiry_date__gte=today).order_by('-date_of_app')
            else:
                matured_jobs = models.JobAdvertisementModel.objects.filter(
                                expiry_date__gte=today).order_by('-date_of_app')
            jobs = []
            for mature in matured_jobs:
                if (mature.expiry_date-today).days<=30:
                    jobs.append(mature)
            result_cnt = len(jobs)

        else:
            if other_url is None and other_url != 'matured-jobs' and other_url != 'remainder-jobs':
                if keyword is None:
                    jobs = models.JobAdvertisementModel.objects.all().order_by('-date_of_app')
                    result_cnt = jobs.count()

        jobs_page = Paginator(jobs, 25)
        page_number = request.GET.get('page')

        try:
            jobs_page = jobs_page.get_page(page_number)
        except PageNotAnInteger:
            jobs_page = jobs_page.page(1)
        except EmptyPage:
            jobs_page = jobs_page.page(jobs_page.num_pages)
        datas = {
            'jobs_page': jobs_page,
            'current_url': current_url,
            'keyword': keyword,
            'result_cnt': result_cnt,
        }
        return render(request, self.template_name, context=datas)

@method_decorator(login_required(login_url="/"), name='dispatch')
class UpdateReviewJob(View):
    def post(self, request, update_review_id):
        update_date_of_review = request.POST.get('update_date_of_review')
        update_no_of_app = request.POST.get('update_no_of_app')
        try:
            update_date_of_review = datetime.strptime(update_date_of_review, '%d/%m/%Y').date()
        except Exception:
            update_date_of_review = None

        try:
            update_job = models.JobAdvertisementModel.objects.get(id=update_review_id)
            update_job.date_of_review = update_date_of_review
            update_job.no_of_app = update_no_of_app
            update_job.save()
            messages.success(request, f'{update_job.job_code} DOR : {update_date_of_review} No.of.App : {update_no_of_app} Updated..')
        except Exception as e:
            messages.error(request, f'Error: {e}')
        return redirect('/job-advertisement/job-list/matured-jobs')

@method_decorator(login_required(login_url="/"), name='dispatch')
class DeleteJobView(View):
    def get(self, request, id):
        try:
            job = models.JobAdvertisementModel.objects.get(id=id)
            job.delete()
        except Exception as e:
            messages.error(request, f'Error: {e}')
        return redirect('/job-advertisement/job-list')

class MyTodoAPIView(View):
    def post(self, request):
        deadline = request.POST.get('deadline')
        description = request.POST.get('description')
        try:
            deadline = datetime.strptime(deadline, '%Y-%m-%dT%H:%M')
            user = User.objects.get(username=request.user.username)
            task = models.MyTodoModel(user=user, deadline=deadline, task=description)
            task.save()
            messages.success(request, f'New Task Added..')
            response_data = {'status': 'success', 'msg': 'New Task Added.'}
        except Exception as e:
            response_data = {'status': 'error', 'msg': 'Error: {e}', }
            messages.error(request, f'Error: {e}')
        return JsonResponse(response_data)

def is_completed_todo_api(request, id):
    if request.method == 'GET':
        task = models.MyTodoModel.objects.get(id=id)
        if task.is_completed:
            task.is_completed = False
            msg = 'false'
        else:
            task.is_completed = True
            msg = 'true'
        task.save()
        response_data = {'msg': msg}
        return JsonResponse(response_data)

def delete_todo_api(request, id):
    if request.method == 'GET':
        try:
            task = models.MyTodoModel.objects.get(id=id)
            task.delete()
            response_data = {'msg': 'success'}
        except Exception as e:
            response_data = {'msg': 'error'}
        return JsonResponse(response_data)

@method_decorator(login_required(login_url="/"), name='dispatch')
class MyTodoView(View):
    template_name = 'todo.html'
    def get(self, request):
        import pytz
        utc=pytz.UTC
        today = datetime.today().replace(tzinfo=utc) + timedelta(hours=5, minutes=30)
        tasks = models.MyTodoModel.objects.filter(user=request.user).order_by('-id', '-is_completed')
        datas = {'tasks': tasks, 'today': today}
        return render(request, self.template_name, context=datas)

def read_job_file_column(ws):
        COLUMN_NAME_FORMAT = ['D.O.APPLICATION', 'JOB CODE', 'MATURITY', 'EXPIRY DATE', 'VACANCIES',
                            'DESIGNATION', 'COMPANY ID', 'DATE OF REVIEW', 'NO OF APP', 'APPLIED EMPLOYEE NAME']
        column = []
        try:
            for col in range(1, 11):
                col_val = ws.cell(row=1, column=col).value
                column.append(str(col_val).upper().strip())
            return column == COLUMN_NAME_FORMAT
        except Exception as e:
            return f'Error: {e}'

def error_table_html(errors):
    html = '''
            <table class="table table-striped custom-table mb-0 datatable">
                <thead>
                    <tr>
                        <th style="text-align:center">#</th>
                        <th style="text-align:center">Errors: Please Fix and Reupload</th>
                    </tr>
                </thead>
                <tbody>
                    {}
                </tbody>
            </table>
            '''
    trows = ''
    for indx, er in enumerate(errors, 1):
        trows += f'''
                    <tr>
                        <td style="text-align:center">{indx}</td>
                        <td style="text-align:center">{er}</td>
                    </tr>
                 '''
    html_table = html.format(trows)
    return html_table

class InsertJobBulkUploadAPI(View):
    def post(self, request):
        filename = request.FILES.get('myFile')
        errors, insert_datas = [], []
        try:
            if not filename.name.endswith('.xlsx'):
                errors.append('File format should be ".xlsx", please change the file format')
                return errors
            wb = openpyxl.load_workbook(filename)
            ws = wb.worksheets[0]
            get_excel_column = read_job_file_column(ws)
            if isinstance(get_excel_column, bool) and not get_excel_column:
                errors.append('"Sheet 1: Column Name" not Matched, please check sample file format')
                return errors
            elif isinstance(get_excel_column, str):
                errors.append(get_excel_column)
                return errors
            try:
                MAX_ROW_LIMIT = 100000
                excel_max_row = ws.max_row + 1
                if excel_max_row > MAX_ROW_LIMIT:
                    errors.append(f'MAX LIMIT: {MAX_ROW_LIMIT}, Excel Rows Exceeds {MAX_ROW_LIMIT}.')
                    return errors
                for row in range(2, excel_max_row):
                    d_o_app = ws.cell(row=row, column=1)
                    job_code = ws.cell(row=row, column=2)
                    maturity = ws.cell(row=row, column=3)
                    expiry_date = ws.cell(row=row, column=4)
                    vacancies = ws.cell(row=row, column=5)
                    designation = ws.cell(row=row, column=6)
                    company_id = ws.cell(row=row, column=7)
                    date_of_review = ws.cell(row=row, column=8)
                    no_of_app = ws.cell(row=row, column=9)
                    applied_emp_name = ws.cell(row=row, column=10)

                    # mandatory field: d_o_app, job_code, maturity, expiry_date, vacancies, designation
                    if d_o_app.value is None:
                        er = f'Row: {row} Field: {d_o_app.coordinate} Mandatory Field.'
                        errors.append(er)
                    if job_code.value is None:
                        er = f'Row: {row} Field: {job_code.coordinate} Mandatory Field.'
                        errors.append(er)
                    if maturity.value is None:
                        er = f'Row: {row} Field: {maturity.coordinate} Mandatory Field.'
                        errors.append(er)
                    if expiry_date.value is None:
                        er = f'Row: {row} Field: {expiry_date.coordinate} Mandatory Field.'
                        errors.append(er)
                    if vacancies.value is None:
                        er = f'Row: {row} Field: {vacancies.coordinate} Mandatory Field.'
                        errors.append(er)
                    if designation.value is None:
                        er = f'Row: {row} Field: {designation.coordinate} Mandatory Field.'
                        errors.append(er)

                    try:
                        if not d_o_app.is_date:
                            d_o_app = datetime.strptime(d_o_app.value, '%d/%m/%Y').date()
                        else:
                            d_o_app = d_o_app.value
                    except:
                        er = f'Row: {row} Field: {d_o_app.coordinate} Wrong Date Format. Fomat should be DD/MM/YYYY'
                        errors.append(er)
                    try:
                        if not maturity.is_date:
                            maturity = datetime.strptime(maturity.value, '%d/%m/%Y').date()
                        else:
                            maturity = maturity.value
                    except:
                        er = f'Row: {row} Field: {maturity.coordinate} Wrong Date Format. Fomat should be DD/MM/YYYY'
                        errors.append(er)
                    try:
                        if not expiry_date.is_date:
                            expiry_date = datetime.strptime(expiry_date.value, '%d/%m/%Y').date()
                        else:
                            expiry_date = expiry_date.value
                    except:
                        er = f'Row: {row} Field: {expiry_date.coordinate} Wrong Date Format. Fomat should be DD/MM/YYYY'
                        errors.append(er)
                    try:
                        if date_of_review.value is not None and not date_of_review.is_date:
                            date_of_review = datetime.strptime(date_of_review.value, '%d/%m/%Y').date()
                        else:
                            date_of_review = date_of_review.value
                    except:
                        er = f'Row: {row} Field: {date_of_review.coordinate} Wrong Date Format. Fomat should be DD/MM/YYYY'
                        errors.append(er)

                    try:
                        job_code = str(job_code.value).strip()
                        try:
                            vacancies = int(str(vacancies.value).strip())
                        except:
                            vacancies = 0
                        designation = str(designation.value).strip()
                        if company_id.value is not None:
                            company_id = str(company_id.value).strip()
                        else:
                            company_id = None
                        if applied_emp_name.value is not None:
                            applied_emp_name = str(applied_emp_name.value).strip()
                        else:
                            applied_emp_name = None
                        try:
                            no_of_app = int(str(no_of_app.value).strip())
                        except:
                            no_of_app = 0
                    except Exception as e:
                        errors.append(f'Error: {e}')

                    if models.JobAdvertisementModel.objects.filter(job_code=job_code):
                        er = f'Row: {row} Field: {date_of_review.coordinate}. JobCode: {job_code} Already Exist!'
                        errors.append(er)
                    insert_datas.append({
                        'd_o_app': d_o_app,
                        'job_code': job_code,
                        'maturity': maturity,
                        'expiry_date': expiry_date,
                        'vacancies': vacancies,
                        'designation': designation,
                        'company_id': company_id,
                        'date_of_review': date_of_review,
                        'no_of_app': no_of_app,
                        'applied_emp_name': applied_emp_name,
                    })
            except Exception as e:
                errors.append(f'Error: {e}')
                return errors
        except Exception as e:
                errors.append(f'Error: {e}')
                return errors
        if not errors:
            cnt = 0
            for insert_data in insert_datas:
                try:
                    company_id = AddCompanyModel.objects.get(company_id=insert_data['company_id'])
                except:
                    company_id = None
                try:
                    job = models.JobAdvertisementModel(date_of_app=insert_data['d_o_app'], maturity_date=insert_data['maturity'],
                            expiry_date=insert_data['expiry_date'], job_code=insert_data['job_code'],
                            range_field=None, designation=insert_data['designation'], comp_name=company_id,
                            no_of_vaccancies=insert_data['vacancies'], date_of_review=insert_data['date_of_review'],
                            no_of_app=insert_data['no_of_app'])
                    job.save()
                    cnt = cnt + 1
                except: pass
            msg = f'{cnt} rows inserted'
            messages.success(request, msg)
            return JsonResponse({'status': 'success', 'msg': msg})

        error_table = error_table_html(errors)
        return JsonResponse({'status': 'error', 'table': error_table})

@method_decorator(login_required(login_url="/"), name='dispatch')
class JobBulkUploadView(View):
    template_name = 'job-bulk-upload.html'
    def get(self, request):
        return render(request, self.template_name)


# Logging View
@method_decorator(login_required(login_url="/"), name='dispatch')
class UserAccessLogging(View):
    def get(self, request):
        from django_db_logger import models
        logs = models.StatusLog.objects.all()
        logs = Paginator(logs, 25)
        page_number = request.GET.get('page')
        try:
            logs = logs.get_page(page_number)
        except PageNotAnInteger:
            logs = logs.page(1)
        except EmptyPage:
            logs = logs.page(logs.num_pages)
        datas = {
            'access_logs': logs,
        }
        return render(request, 'access-log.html', context=datas)

@method_decorator(login_required(login_url="/"), name='dispatch')
class ActiveUsersLogView(View):
    def get_all_logged_in_users(self):
        users = models.RealtimeUserActiveModel.objects.all()
        active_users = []
        try:
            for usr in users:
                current_date = datetime.today()
                last_active_time = usr.active_date.replace(tzinfo=None)
                active_diff = current_date - last_active_time
                expected_seconds = timedelta(seconds=10)
                if expected_seconds > active_diff:
                    active_users.append(usr)
        except: pass
        return active_users

    def get(self, request):
        users = self.get_all_logged_in_users()
        return render(request, 'active-users.html', context={'users': users})

class RealtimeUserOnlineAPI(View):
    def post(self, request):
        try:
            user = User.objects.get(username=request.user)

            realtime = models.RealtimeUserActiveModel.objects.filter(user=user)
            if realtime:
                tday = datetime.today()
                realtime = realtime.first()
                realtime.active_date = tday
                realtime.save()
            else:
                tday = datetime.today()
                realtime = models.RealtimeUserActiveModel(user=user, active_date=tday)
                realtime.save()
        except:
            pass
        response_data = {'status': 'success'}
        return JsonResponse(response_data)

class ActiveUserPageUpdateAPI(View):
    def get(self, request):
        users = models.RealtimeUserActiveModel.objects.all()
        active_users = []
        try:
            for usr in users:
                current_date = datetime.today()
                last_active_time = usr.active_date.replace(tzinfo=None)
                active_diff = current_date - last_active_time
                expected_seconds = timedelta(seconds=10)
                if expected_seconds > active_diff:
                    active_users.append(usr)
        except: pass
        html = ''
        for usr in active_users:
            auser = usr.user.username
            active_date = (usr.active_date + timedelta(hours=5, minutes=30)).strftime('%B %d, %Y, %I:%M %p')
            trow = '''
                    <tr>
                        <td style="text-align:center">{}</td>
                        <td style="text-align:center">{}</td>
                        <td style="text-align:center"><p class="badge bg-inverse-success">Online</p></td>
                    </tr>
                    '''.format(auser, active_date)
            html += trow

        response_data = {'status': 'success', 'trows': html}
        return JsonResponse(response_data)